"""ExcelDocument - 双引擎 Excel 文档管理器

实现 pandas (分析引擎) + openpyxl (操作引擎) 的读写分离架构。
"""

import re
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    Color, NamedStyle
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1


class ChangeType(Enum):
    """变更类型"""
    CELL_VALUE = "cell_value"
    CELL_FORMULA = "cell_formula"
    INSERT_ROWS = "insert_rows"
    DELETE_ROWS = "delete_rows"
    INSERT_COLS = "insert_cols"
    DELETE_COLS = "delete_cols"
    STYLE = "style"


@dataclass
class Change:
    """变更记录"""
    change_type: ChangeType
    sheet_name: str
    location: str  # 单元格地址或范围
    old_value: Any = None
    new_value: Any = None
    timestamp: datetime = field(default_factory=datetime.now)


class ExcelDocument:
    """双引擎 Excel 文档管理器

    特性:
    - pandas DataFrame 用于快速查询和分析
    - openpyxl Workbook 用于写入、公式和格式操作
    - 惰性同步：写入后标记 dirty，下次读取时按需同步
    - 变更追踪：记录所有写入操作
    """

    def __init__(self):
        # 双引擎
        self._workbook: Optional[Workbook] = None
        self._dataframes: Dict[str, pd.DataFrame] = {}  # sheet_name -> DataFrame

        # 文件信息
        self._file_path: Optional[str] = None
        self._active_sheet: Optional[str] = None
        self._all_sheets: List[str] = []
        self._is_csv: bool = False  # 是否为 CSV 文件

        # 状态追踪
        self._is_dirty: bool = False
        self._dirty_sheets: set = set()  # 需要同步的工作表
        self._change_log: List[Change] = []
        self._data_version: int = 0  # 数据版本号，用于缓存失效

    # ==================== 属性访问 ====================

    @property
    def is_loaded(self) -> bool:
        """是否已加载文件"""
        return self._workbook is not None

    @property
    def is_dirty(self) -> bool:
        """是否有未保存的修改"""
        return self._is_dirty

    @property
    def file_path(self) -> Optional[str]:
        """文件路径"""
        return self._file_path

    @property
    def active_sheet(self) -> Optional[str]:
        """当前活跃工作表名称"""
        return self._active_sheet

    @property
    def all_sheets(self) -> List[str]:
        """所有工作表名称"""
        return self._all_sheets.copy()

    @property
    def data_version(self) -> int:
        """数据版本号"""
        return self._data_version

    @property
    def change_log(self) -> List[Change]:
        """变更日志"""
        return self._change_log.copy()

    # ==================== 引擎访问 ====================

    @property
    def dataframe(self) -> pd.DataFrame:
        """获取当前工作表的 DataFrame (分析引擎)

        如果工作表被修改过，会先同步数据。
        """
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        sheet = self._active_sheet

        # 如果该表被修改过，先同步
        if sheet in self._dirty_sheets:
            self._sync_sheet_to_df(sheet)

        if sheet not in self._dataframes:
            self._load_sheet_to_df(sheet)

        return self._dataframes[sheet]

    @property
    def workbook(self) -> Workbook:
        """获取 Workbook (操作引擎)"""
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")
        return self._workbook

    @property
    def worksheet(self) -> Worksheet:
        """获取当前活跃工作表"""
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")
        return self._workbook[self._active_sheet]

    def get_read_engine(self) -> pd.DataFrame:
        """获取分析引擎 (DataFrame)"""
        return self.dataframe

    def get_write_engine(self) -> Worksheet:
        """获取操作引擎 (Worksheet)"""
        return self.worksheet

    # ==================== 加载与保存 ====================

    def load(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """加载 Excel 或 CSV 文件

        Args:
            file_path: Excel/CSV 文件路径
            sheet_name: 工作表名称，默认加载第一个（CSV 文件忽略此参数）

        Returns:
            文件结构信息
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        suffix = path.suffix.lower()
        if suffix not in ['.xlsx', '.xlsm', '.csv']:
            raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx, .xlsm, .csv")

        self._file_path = str(path.absolute())
        self._is_csv = suffix == '.csv'

        if suffix == '.csv':
            # CSV 文件处理：创建一个新的 Workbook 并导入数据
            df = pd.read_csv(file_path, encoding='utf-8')
            self._workbook = Workbook()
            ws = self._workbook.active
            ws.title = 'Sheet1'

            # 将 DataFrame 数据写入 Worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            self._all_sheets = ['Sheet1']
            self._active_sheet = 'Sheet1'
        else:
            # Excel 文件处理
            self._workbook = load_workbook(file_path, data_only=False)
            self._all_sheets = self._workbook.sheetnames

            # 确定活跃工作表
            if sheet_name is None:
                self._active_sheet = self._all_sheets[0]
            elif sheet_name not in self._all_sheets:
                raise ValueError(f"工作表 '{sheet_name}' 不存在，可用: {self._all_sheets}")
            else:
                self._active_sheet = sheet_name

        # 初始加载当前工作表到 DataFrame
        self._load_sheet_to_df(self._active_sheet)

        # 重置状态
        self._is_dirty = False
        self._dirty_sheets.clear()
        self._change_log.clear()
        self._data_version = 0

        return self.get_structure()

    def save(self, file_path: Optional[str] = None) -> str:
        """保存文件

        Args:
            file_path: 保存路径，默认覆盖原文件
                      如果原文件是 CSV 且未指定路径，将保存为同名 .xlsx 文件
                      如果指定 .csv 后缀，将导出为 CSV 格式

        Returns:
            保存的文件路径
        """
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        save_path = file_path or self._file_path
        if not save_path:
            raise ValueError("未指定保存路径")

        save_path = Path(save_path)
        suffix = save_path.suffix.lower()

        # 如果原文件是 CSV 且没有指定新路径，默认保存为 xlsx
        if self._is_csv and file_path is None:
            save_path = save_path.with_suffix('.xlsx')
            suffix = '.xlsx'

        if suffix == '.csv':
            # 保存为 CSV 格式（只保存当前活跃工作表）
            df = self.dataframe
            df.to_csv(str(save_path), index=False, encoding='utf-8')
        else:
            # 保存为 Excel 格式
            self._workbook.save(str(save_path))

        # 如果是另存为，更新文件路径和状态
        if file_path:
            self._file_path = str(save_path.absolute())
            self._is_csv = suffix == '.csv'

        # 重置 dirty 状态
        self._is_dirty = False
        self._dirty_sheets.clear()

        return str(save_path)

    def save_as(self, file_path: str) -> str:
        """另存为

        Args:
            file_path: 新文件路径

        Returns:
            保存的文件路径
        """
        return self.save(file_path)

    def create_new(self) -> None:
        """创建新的空白工作簿"""
        self._workbook = Workbook()
        self._file_path = None
        self._all_sheets = self._workbook.sheetnames
        self._active_sheet = self._all_sheets[0]
        self._dataframes = {}
        self._is_dirty = True
        self._dirty_sheets.clear()
        self._change_log.clear()
        self._data_version = 0

    # ==================== 工作表操作 ====================

    def switch_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """切换到指定工作表

        Args:
            sheet_name: 工作表名称

        Returns:
            切换后的结构信息
        """
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        if sheet_name not in self._all_sheets:
            raise ValueError(f"工作表 '{sheet_name}' 不存在，可用: {self._all_sheets}")

        self._active_sheet = sheet_name

        # 按需加载 DataFrame
        if sheet_name not in self._dataframes:
            self._load_sheet_to_df(sheet_name)

        return self.get_structure()

    def create_sheet(self, name: str, index: Optional[int] = None) -> str:
        """创建新工作表

        Args:
            name: 工作表名称
            index: 插入位置，默认在末尾

        Returns:
            新工作表名称
        """
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        if name in self._all_sheets:
            raise ValueError(f"工作表 '{name}' 已存在")

        self._workbook.create_sheet(name, index)
        self._all_sheets = self._workbook.sheetnames
        self._mark_dirty()

        return name

    def delete_sheet(self, name: str) -> bool:
        """删除工作表

        Args:
            name: 工作表名称

        Returns:
            是否删除成功
        """
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        if name not in self._all_sheets:
            return False

        if len(self._all_sheets) <= 1:
            raise ValueError("不能删除最后一个工作表")

        del self._workbook[name]
        self._all_sheets = self._workbook.sheetnames

        # 清理缓存
        if name in self._dataframes:
            del self._dataframes[name]
        if name in self._dirty_sheets:
            self._dirty_sheets.remove(name)

        # 如果删除的是活跃表，切换到第一个
        if self._active_sheet == name:
            self._active_sheet = self._all_sheets[0]

        self._mark_dirty()
        return True

    def rename_sheet(self, old_name: str, new_name: str) -> bool:
        """重命名工作表

        Args:
            old_name: 原名称
            new_name: 新名称

        Returns:
            是否重命名成功
        """
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        if old_name not in self._all_sheets:
            return False

        if new_name in self._all_sheets:
            raise ValueError(f"工作表 '{new_name}' 已存在")

        self._workbook[old_name].title = new_name
        self._all_sheets = self._workbook.sheetnames

        # 更新缓存
        if old_name in self._dataframes:
            self._dataframes[new_name] = self._dataframes.pop(old_name)
        if old_name in self._dirty_sheets:
            self._dirty_sheets.remove(old_name)
            self._dirty_sheets.add(new_name)

        # 更新活跃表名
        if self._active_sheet == old_name:
            self._active_sheet = new_name

        self._mark_dirty()
        return True

    # ==================== 单元格读写 ====================

    def read_cell(self, cell: str, sheet: Optional[str] = None) -> Any:
        """读取单元格值

        Args:
            cell: 单元格地址，如 "A1"
            sheet: 工作表名称，默认当前表

        Returns:
            单元格值
        """
        ws = self._get_worksheet(sheet)
        return ws[cell].value

    def write_cell(
        self,
        cell: str,
        value: Any,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """写入单元格

        Args:
            cell: 单元格地址，如 "A1"
            value: 写入的值
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        old_value = ws[cell].value
        ws[cell].value = value

        # 记录变更
        self._record_change(
            ChangeType.CELL_VALUE,
            sheet,
            cell,
            old_value,
            value
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell": cell,
            "sheet": sheet,
            "old_value": old_value,
            "new_value": value
        }

    def read_range(
        self,
        start_cell: str,
        end_cell: str,
        sheet: Optional[str] = None
    ) -> List[List[Any]]:
        """读取单元格范围

        Args:
            start_cell: 起始单元格，如 "A1"
            end_cell: 结束单元格，如 "C10"
            sheet: 工作表名称，默认当前表

        Returns:
            二维数据数组
        """
        ws = self._get_worksheet(sheet)

        data = []
        for row in ws[f"{start_cell}:{end_cell}"]:
            row_data = [cell.value for cell in row]
            data.append(row_data)

        return data

    def write_range(
        self,
        start_cell: str,
        data: List[List[Any]],
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """批量写入数据

        Args:
            start_cell: 起始单元格，如 "A1"
            data: 二维数据数组
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        start_row, start_col = self._parse_cell_address(start_cell)

        cells_written = 0
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx)
                cell.value = value
                cells_written += 1

        # 计算结束单元格
        end_row = start_row + len(data) - 1
        end_col = start_col + (max(len(row) for row in data) if data else 0) - 1
        end_cell = f"{get_column_letter(end_col)}{end_row}"

        # 记录变更
        self._record_change(
            ChangeType.CELL_VALUE,
            sheet,
            f"{start_cell}:{end_cell}",
            None,
            f"写入 {len(data)} 行数据"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "start_cell": start_cell,
            "end_cell": end_cell,
            "sheet": sheet,
            "rows_written": len(data),
            "cells_written": cells_written
        }

    # ==================== 公式操作 ====================

    def read_formula(self, cell: str, sheet: Optional[str] = None) -> Optional[str]:
        """读取单元格公式

        Args:
            cell: 单元格地址
            sheet: 工作表名称，默认当前表

        Returns:
            公式字符串（不含 =），如果不是公式则返回 None
        """
        ws = self._get_worksheet(sheet)
        cell_obj = ws[cell]

        if cell_obj.data_type == 'f' or (
            isinstance(cell_obj.value, str) and cell_obj.value.startswith('=')
        ):
            value = cell_obj.value
            return value[1:] if value.startswith('=') else value

        return None

    def write_formula(
        self,
        cell: str,
        formula: str,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """写入公式

        Args:
            cell: 单元格地址
            formula: 公式字符串，如 "SUM(A1:A10)" 或 "=SUM(A1:A10)"
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        # 确保公式以 = 开头
        if not formula.startswith('='):
            formula = f"={formula}"

        old_value = ws[cell].value
        ws[cell].value = formula

        # 记录变更
        self._record_change(
            ChangeType.CELL_FORMULA,
            sheet,
            cell,
            old_value,
            formula
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell": cell,
            "sheet": sheet,
            "formula": formula,
            "note": "公式将在 Excel 中打开时计算"
        }

    def list_formulas(self, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
        """列出工作表中的所有公式

        Args:
            sheet: 工作表名称，默认当前表

        Returns:
            公式列表
        """
        ws = self._get_worksheet(sheet)
        formulas = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' or (
                    isinstance(cell.value, str) and
                    cell.value and
                    cell.value.startswith('=')
                ):
                    formulas.append({
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "row": cell.row,
                        "column": get_column_letter(cell.column)
                    })

        return formulas

    # ==================== 行列操作 ====================

    def insert_rows(
        self,
        row: int,
        count: int = 1,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """插入行

        Args:
            row: 在此行之前插入
            count: 插入行数
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        ws.insert_rows(row, count)

        # 记录变更
        self._record_change(
            ChangeType.INSERT_ROWS,
            sheet,
            f"行 {row}",
            None,
            f"插入 {count} 行"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "sheet": sheet,
            "row": row,
            "count": count,
            "operation": "insert_rows"
        }

    def delete_rows(
        self,
        start_row: int,
        end_row: Optional[int] = None,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """删除行

        Args:
            start_row: 起始行号
            end_row: 结束行号（含），默认只删除起始行
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        if end_row is None:
            end_row = start_row

        count = end_row - start_row + 1
        ws.delete_rows(start_row, count)

        # 记录变更
        self._record_change(
            ChangeType.DELETE_ROWS,
            sheet,
            f"行 {start_row}-{end_row}",
            f"删除 {count} 行",
            None
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "sheet": sheet,
            "start_row": start_row,
            "end_row": end_row,
            "count": count,
            "operation": "delete_rows"
        }

    def insert_cols(
        self,
        col: int,
        count: int = 1,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """插入列

        Args:
            col: 在此列之前插入（数字，1 = A）
            count: 插入列数
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        ws.insert_cols(col, count)

        # 记录变更
        self._record_change(
            ChangeType.INSERT_COLS,
            sheet,
            f"列 {get_column_letter(col)}",
            None,
            f"插入 {count} 列"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "sheet": sheet,
            "column": get_column_letter(col),
            "count": count,
            "operation": "insert_cols"
        }

    def delete_cols(
        self,
        start_col: int,
        end_col: Optional[int] = None,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """删除列

        Args:
            start_col: 起始列号（数字）
            end_col: 结束列号（数字），默认只删除起始列
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        if end_col is None:
            end_col = start_col

        count = end_col - start_col + 1
        ws.delete_cols(start_col, count)

        # 记录变更
        self._record_change(
            ChangeType.DELETE_COLS,
            sheet,
            f"列 {get_column_letter(start_col)}-{get_column_letter(end_col)}",
            f"删除 {count} 列",
            None
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "sheet": sheet,
            "start_col": get_column_letter(start_col),
            "end_col": get_column_letter(end_col),
            "count": count,
            "operation": "delete_cols"
        }

    # ==================== 格式化操作 ====================

    def set_font(
        self,
        cell_range: str,
        name: Optional[str] = None,
        size: Optional[int] = None,
        bold: Optional[bool] = None,
        italic: Optional[bool] = None,
        underline: Optional[str] = None,
        color: Optional[str] = None,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """设置字体样式

        Args:
            cell_range: 单元格或范围，如 "A1" 或 "A1:C10"
            name: 字体名称，如 "Arial", "宋体"
            size: 字号
            bold: 是否加粗
            italic: 是否斜体
            underline: 下划线类型 ("single", "double", "singleAccounting", "doubleAccounting")
            color: 字体颜色 (十六进制，如 "FF0000" 表示红色)
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        # 创建字体对象
        font_kwargs = {}
        if name is not None:
            font_kwargs["name"] = name
        if size is not None:
            font_kwargs["sz"] = size
        if bold is not None:
            font_kwargs["bold"] = bold
        if italic is not None:
            font_kwargs["italic"] = italic
        if underline is not None:
            font_kwargs["underline"] = underline
        if color is not None:
            font_kwargs["color"] = Color(rgb=color)

        if not font_kwargs:
            return {"success": False, "error": "未指定任何字体属性"}

        cells_modified = self._apply_style_to_range(ws, cell_range, font=Font(**font_kwargs))

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            f"设置字体: {font_kwargs}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "cells_modified": cells_modified,
            "font_settings": {k: str(v) for k, v in font_kwargs.items()}
        }

    def set_fill(
        self,
        cell_range: str,
        color: str,
        fill_type: str = "solid",
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """设置单元格背景色

        Args:
            cell_range: 单元格或范围，如 "A1" 或 "A1:C10"
            color: 背景颜色 (十六进制，如 "FFFF00" 表示黄色)
            fill_type: 填充类型 ("solid", "darkGray", "mediumGray", "lightGray", "gray125", "gray0625")
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)
        cells_modified = self._apply_style_to_range(ws, cell_range, fill=fill)

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            f"设置背景色: {color}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "cells_modified": cells_modified,
            "color": color,
            "fill_type": fill_type
        }

    def set_alignment(
        self,
        cell_range: str,
        horizontal: Optional[str] = None,
        vertical: Optional[str] = None,
        wrap_text: Optional[bool] = None,
        text_rotation: Optional[int] = None,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """设置单元格对齐方式

        Args:
            cell_range: 单元格或范围
            horizontal: 水平对齐 ("left", "center", "right", "justify", "distributed")
            vertical: 垂直对齐 ("top", "center", "bottom", "justify", "distributed")
            wrap_text: 是否自动换行
            text_rotation: 文本旋转角度 (0-180)
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        align_kwargs = {}
        if horizontal is not None:
            align_kwargs["horizontal"] = horizontal
        if vertical is not None:
            align_kwargs["vertical"] = vertical
        if wrap_text is not None:
            align_kwargs["wrap_text"] = wrap_text
        if text_rotation is not None:
            align_kwargs["text_rotation"] = text_rotation

        if not align_kwargs:
            return {"success": False, "error": "未指定任何对齐属性"}

        alignment = Alignment(**align_kwargs)
        cells_modified = self._apply_style_to_range(ws, cell_range, alignment=alignment)

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            f"设置对齐: {align_kwargs}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "cells_modified": cells_modified,
            "alignment_settings": align_kwargs
        }

    def set_border(
        self,
        cell_range: str,
        style: str = "thin",
        color: str = "000000",
        left: bool = True,
        right: bool = True,
        top: bool = True,
        bottom: bool = True,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """设置单元格边框

        Args:
            cell_range: 单元格或范围
            style: 边框样式 ("thin", "medium", "thick", "double", "dotted", "dashed")
            color: 边框颜色 (十六进制)
            left: 是否设置左边框
            right: 是否设置右边框
            top: 是否设置上边框
            bottom: 是否设置下边框
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        side = Side(style=style, color=Color(rgb=color))
        border = Border(
            left=side if left else None,
            right=side if right else None,
            top=side if top else None,
            bottom=side if bottom else None
        )
        cells_modified = self._apply_style_to_range(ws, cell_range, border=border)

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            f"设置边框: style={style}, color={color}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "cells_modified": cells_modified,
            "border_style": style,
            "border_color": color
        }

    def set_number_format(
        self,
        cell_range: str,
        format_code: str,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """设置数字格式

        Args:
            cell_range: 单元格或范围
            format_code: 数字格式代码，如:
                - "#,##0" 千分位整数
                - "#,##0.00" 千分位两位小数
                - "0.00%" 百分比
                - "yyyy-mm-dd" 日期
                - "¥#,##0.00" 货币
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        cells_modified = self._apply_style_to_range(ws, cell_range, number_format=format_code)

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            f"设置数字格式: {format_code}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "cells_modified": cells_modified,
            "format_code": format_code
        }

    def set_cell_style(
        self,
        cell_range: str,
        font_name: Optional[str] = None,
        font_size: Optional[int] = None,
        font_bold: Optional[bool] = None,
        font_italic: Optional[bool] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        horizontal: Optional[str] = None,
        vertical: Optional[str] = None,
        border_style: Optional[str] = None,
        border_color: Optional[str] = None,
        number_format: Optional[str] = None,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """综合设置单元格样式

        一次性设置多种样式属性，更高效。

        Args:
            cell_range: 单元格或范围
            font_name: 字体名称
            font_size: 字号
            font_bold: 是否加粗
            font_italic: 是否斜体
            font_color: 字体颜色
            bg_color: 背景颜色
            horizontal: 水平对齐
            vertical: 垂直对齐
            border_style: 边框样式
            border_color: 边框颜色
            number_format: 数字格式
            sheet: 工作表名称

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        style_kwargs = {}
        settings_applied = []

        # 字体
        font_kwargs = {}
        if font_name is not None:
            font_kwargs["name"] = font_name
        if font_size is not None:
            font_kwargs["sz"] = font_size
        if font_bold is not None:
            font_kwargs["bold"] = font_bold
        if font_italic is not None:
            font_kwargs["italic"] = font_italic
        if font_color is not None:
            font_kwargs["color"] = Color(rgb=font_color)
        if font_kwargs:
            style_kwargs["font"] = Font(**font_kwargs)
            settings_applied.append("字体")

        # 背景色
        if bg_color is not None:
            style_kwargs["fill"] = PatternFill(
                start_color=bg_color,
                end_color=bg_color,
                fill_type="solid"
            )
            settings_applied.append("背景色")

        # 对齐
        align_kwargs = {}
        if horizontal is not None:
            align_kwargs["horizontal"] = horizontal
        if vertical is not None:
            align_kwargs["vertical"] = vertical
        if align_kwargs:
            style_kwargs["alignment"] = Alignment(**align_kwargs)
            settings_applied.append("对齐")

        # 边框
        if border_style is not None:
            side = Side(
                style=border_style,
                color=Color(rgb=border_color or "000000")
            )
            style_kwargs["border"] = Border(left=side, right=side, top=side, bottom=side)
            settings_applied.append("边框")

        # 数字格式
        if number_format is not None:
            style_kwargs["number_format"] = number_format
            settings_applied.append("数字格式")

        if not style_kwargs:
            return {"success": False, "error": "未指定任何样式属性"}

        cells_modified = self._apply_style_to_range(ws, cell_range, **style_kwargs)

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            f"设置样式: {', '.join(settings_applied)}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "cells_modified": cells_modified,
            "settings_applied": settings_applied
        }

    def merge_cells(
        self,
        cell_range: str,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """合并单元格

        Args:
            cell_range: 要合并的范围，如 "A1:C3"
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        ws.merge_cells(cell_range)

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            "合并单元格"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "operation": "merge_cells"
        }

    def unmerge_cells(
        self,
        cell_range: str,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """取消合并单元格

        Args:
            cell_range: 要取消合并的范围，如 "A1:C3"
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        ws.unmerge_cells(cell_range)

        self._record_change(
            ChangeType.STYLE,
            sheet,
            cell_range,
            None,
            "取消合并单元格"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "cell_range": cell_range,
            "sheet": sheet,
            "operation": "unmerge_cells"
        }

    def set_column_width(
        self,
        column: str,
        width: float,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """设置列宽

        Args:
            column: 列标识，如 "A" 或 "B"
            width: 列宽（字符数）
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        ws.column_dimensions[column.upper()].width = width

        self._record_change(
            ChangeType.STYLE,
            sheet,
            f"列 {column}",
            None,
            f"设置列宽: {width}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "column": column.upper(),
            "sheet": sheet,
            "width": width
        }

    def set_row_height(
        self,
        row: int,
        height: float,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """设置行高

        Args:
            row: 行号
            height: 行高（磅）
            sheet: 工作表名称，默认当前表

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        ws.row_dimensions[row].height = height

        self._record_change(
            ChangeType.STYLE,
            sheet,
            f"行 {row}",
            None,
            f"设置行高: {height}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "row": row,
            "sheet": sheet,
            "height": height
        }

    def auto_fit_column(
        self,
        column: str,
        sheet: Optional[str] = None,
        min_width: float = 8.0,
        max_width: float = 50.0
    ) -> Dict[str, Any]:
        """自动调整列宽以适应内容

        Args:
            column: 列标识，如 "A"
            sheet: 工作表名称，默认当前表
            min_width: 最小宽度
            max_width: 最大宽度

        Returns:
            操作结果
        """
        sheet = sheet or self._active_sheet
        ws = self._get_worksheet(sheet)

        col_idx = column_index_from_string(column.upper())
        max_length = 0

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # 估算字符宽度（中文字符算 2 个宽度）
                    cell_str = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                    max_length = max(max_length, length)

        # 计算宽度（加一些 padding）
        width = min(max(max_length * 1.1, min_width), max_width)
        ws.column_dimensions[column.upper()].width = width

        self._record_change(
            ChangeType.STYLE,
            sheet,
            f"列 {column}",
            None,
            f"自动调整列宽: {width:.1f}"
        )
        self._mark_dirty(sheet)

        return {
            "success": True,
            "column": column.upper(),
            "sheet": sheet,
            "width": round(width, 1)
        }

    def _apply_style_to_range(
        self,
        ws: Worksheet,
        cell_range: str,
        font: Optional[Font] = None,
        fill: Optional[PatternFill] = None,
        alignment: Optional[Alignment] = None,
        border: Optional[Border] = None,
        number_format: Optional[str] = None
    ) -> int:
        """应用样式到单元格范围

        Args:
            ws: 工作表
            cell_range: 单元格范围
            font: 字体
            fill: 填充
            alignment: 对齐
            border: 边框
            number_format: 数字格式

        Returns:
            修改的单元格数量
        """
        cells_modified = 0

        # 判断是单个单元格还是范围
        if ":" in cell_range:
            # 范围
            for row in ws[cell_range]:
                for cell in row:
                    self._apply_style_to_cell(
                        cell, font, fill, alignment, border, number_format
                    )
                    cells_modified += 1
        else:
            # 单个单元格
            self._apply_style_to_cell(
                ws[cell_range], font, fill, alignment, border, number_format
            )
            cells_modified = 1

        return cells_modified

    def _apply_style_to_cell(
        self,
        cell: Cell,
        font: Optional[Font] = None,
        fill: Optional[PatternFill] = None,
        alignment: Optional[Alignment] = None,
        border: Optional[Border] = None,
        number_format: Optional[str] = None
    ) -> None:
        """应用样式到单个单元格"""
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if alignment is not None:
            cell.alignment = alignment
        if border is not None:
            cell.border = border
        if number_format is not None:
            cell.number_format = number_format

    # ==================== 数据同步 ====================

    def sync_workbook_to_df(self, sheet: Optional[str] = None) -> None:
        """将 Workbook 数据同步到 DataFrame

        Args:
            sheet: 工作表名称，默认同步所有 dirty 的表
        """
        if sheet:
            self._sync_sheet_to_df(sheet)
            if sheet in self._dirty_sheets:
                self._dirty_sheets.remove(sheet)
        else:
            # 同步所有 dirty 的表
            for s in list(self._dirty_sheets):
                self._sync_sheet_to_df(s)
            self._dirty_sheets.clear()

    def sync_df_to_workbook(self, sheet: Optional[str] = None) -> None:
        """将 DataFrame 数据同步到 Workbook（谨慎使用，会覆盖公式）

        Args:
            sheet: 工作表名称，默认当前表
        """
        sheet = sheet or self._active_sheet
        if sheet not in self._dataframes:
            return

        df = self._dataframes[sheet]
        ws = self._workbook[sheet]

        # 清空工作表
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # 写入表头
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                # 处理 pandas 特殊类型
                if pd.isna(value):
                    value = None
                ws.cell(row=row_idx, column=col_idx, value=value)

        self._mark_dirty(sheet)

    # ==================== 结构信息 ====================

    def get_structure(self) -> Dict[str, Any]:
        """获取 Excel 结构信息"""
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        df = self.dataframe

        # 列信息
        columns_info = []
        for col in df.columns:
            col_data = df[col]
            columns_info.append({
                "name": str(col),
                "dtype": str(col_data.dtype),
                "non_null_count": int(col_data.count()),
                "null_count": int(col_data.isna().sum()),
            })

        return {
            "file_path": self._file_path,
            "sheet_name": self._active_sheet,
            "all_sheets": self._all_sheets,
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "columns": columns_info,
            "is_dirty": self._is_dirty,
            "data_version": self._data_version,
        }

    def get_preview(self, n_rows: int = 10) -> Dict[str, Any]:
        """获取数据预览"""
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        df = self.dataframe
        preview_df = df.head(n_rows)

        return {
            "columns": list(df.columns),
            "data": preview_df.to_dict(orient="records"),
            "preview_rows": len(preview_df),
            "total_rows": len(df),
        }

    def get_summary(self) -> str:
        """获取 Excel 摘要信息（用于 Agent 上下文）"""
        if not self.is_loaded:
            return "未加载 Excel 文件"

        structure = self.get_structure()
        preview = self.get_preview()

        dirty_mark = " ⚠️ *有未保存的修改*" if self._is_dirty else ""

        lines = [
            f"📊 **已加载 Excel 文件**: {structure['file_path']}{dirty_mark}",
            f"📋 **当前工作表**: {structure['sheet_name']}",
            f"📑 **所有工作表**: {', '.join(structure['all_sheets'])}",
            f"📏 **数据规模**: {structure['total_rows']} 行 × {structure['total_columns']} 列",
            "",
            "**列信息**:",
        ]

        for col in structure['columns']:
            lines.append(f"  - `{col['name']}` ({col['dtype']}): {col['non_null_count']} 非空值")

        lines.append("")
        lines.append(f"**前 {preview['preview_rows']} 行数据预览**:")

        # 简单表格格式
        if preview['data']:
            headers = preview['columns']
            lines.append("| " + " | ".join(str(h) for h in headers) + " |")
            lines.append("| " + " | ".join("---" for _ in headers) + " |")
            for row in preview['data']:
                values = [str(row.get(h, ""))[:20] for h in headers]
                lines.append("| " + " | ".join(values) + " |")

        return "\n".join(lines)

    # ==================== 内部方法 ====================

    def _get_worksheet(self, sheet: Optional[str] = None) -> Worksheet:
        """获取指定工作表"""
        if not self.is_loaded:
            raise ValueError("未加载 Excel 文件")

        sheet = sheet or self._active_sheet
        if sheet not in self._all_sheets:
            raise ValueError(f"工作表 '{sheet}' 不存在")

        return self._workbook[sheet]

    def _load_sheet_to_df(self, sheet_name: str) -> None:
        """从 Workbook 加载工作表到 DataFrame"""
        ws = self._workbook[sheet_name]

        # 使用 openpyxl 数据构建 DataFrame
        data = []
        headers = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = list(row)
            else:
                data.append(row)

        if headers:
            self._dataframes[sheet_name] = pd.DataFrame(data, columns=headers)
        else:
            self._dataframes[sheet_name] = pd.DataFrame()

    def _sync_sheet_to_df(self, sheet_name: str) -> None:
        """同步单个工作表到 DataFrame"""
        self._load_sheet_to_df(sheet_name)

    def _mark_dirty(self, sheet: Optional[str] = None) -> None:
        """标记为已修改"""
        self._is_dirty = True
        self._data_version += 1
        if sheet:
            self._dirty_sheets.add(sheet)

    def _record_change(
        self,
        change_type: ChangeType,
        sheet: str,
        location: str,
        old_value: Any = None,
        new_value: Any = None
    ) -> None:
        """记录变更"""
        self._change_log.append(Change(
            change_type=change_type,
            sheet_name=sheet,
            location=location,
            old_value=old_value,
            new_value=new_value
        ))

    def _parse_cell_address(self, cell: str) -> Tuple[int, int]:
        """解析单元格地址为 (行, 列)"""
        match = re.match(r'^([A-Z]+)(\d+)$', cell.upper())
        if not match:
            raise ValueError(f"无效的单元格地址: {cell}")

        col_str, row_str = match.groups()
        col = column_index_from_string(col_str)
        row = int(row_str)

        return row, col

    def __repr__(self) -> str:
        if not self.is_loaded:
            return "ExcelDocument(未加载)"
        return f"ExcelDocument({self._file_path}, sheet='{self._active_sheet}', dirty={self._is_dirty})"
